from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "Workflows" / "real-device-acceptance.xlsx"

headers = [
    "工步号", "设备分类", "设备名称", "通道/端口号", "测试项名称", "动作指令",
    "配置参数1", "配置参数2", "配置参数3", "下限值", "上限值", "单位",
    "超时(ms)", "失败处理", "备注", "解码配置"
]

steps = [
    [1, "DoIP", "DoIP_ECU", "TCP:13400", "DiagnosticSessionControl 10 01", "send_uds",
     "host=192.168.0.100,port=13400,sourceAddress=0x102D,targetAddress=0x1008,activationType=0x00,udsPayload=10 01,expectedPrefix=50 01,connectionMode=session",
     "", "", "", "", "hex", 5000, "重试1次", "替换 ECU IP、逻辑地址和 UDS 期望响应", ""],
    [2, "Modbus", "Modbus_PLC", "TCP:502", "Read Holding Register 0", "read",
     "transport=tcp,host=192.168.0.101,port=502,slaveId=1,functionCode=3,startAddr=0,quantity=1,valueType=uint16,valueOffset=0,valueLength=2",
     "", "", 0, 65535, "counts", 5000, "重试1次", "替换寄存器地址、从站号和工程限值",
     json.dumps([{"name": "register_value", "offset": 0, "length": 2, "type": "uint16", "endian": "big", "unit": "counts"}], ensure_ascii=False)],
    [3, "SCPI", "SCPI_DMM", "TCP:5025", "Instrument Identification", "*IDN?",
     "scpi.idn", ".+", "5000", "", "", "text", 5000, "停机", "确认仪器返回非空身份信息", ""],
    [4, "SCPI", "SCPI_DMM", "TCP:5025", "DC Voltage", "MEAS:VOLT:DC?",
     "scpi.voltage", "[-+0-9.eE]+", "5000", 0, 1000, "V", 5000, "停机", "将上下限改为 DUT 规格；默认范围仅用于连通性验证",
     json.dumps([{"name": "scpi_value", "type": "double", "unit": "V", "trim": True}], ensure_ascii=False)],
]

thin = Side(style="thin", color="9AA7B8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1A3050")

wb = Workbook()
ws = wb.active
ws.title = "测试工步表"
for col, value in enumerate(headers, 1):
    cell = ws.cell(1, col, value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border
for row, values in enumerate(steps, 2):
    for col, value in enumerate(values, 1):
        cell = ws.cell(row, col, value if value != "" else None)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
for col, width in enumerate([8, 12, 18, 14, 28, 22, 72, 20, 12, 12, 12, 10, 10, 12, 42, 72], 1):
    ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = width
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:P{len(steps) + 1}"

mapping = wb.create_sheet("设备通道映射表")
mapping_headers = ["设备编号", "设备分类", "设备名称", "通道号", "连接方式", "主地址", "子端口", "从地址", "设备型号", "校准有效期", "数据位", "停止位", "校验位", "启用", "备注"]
for col, value in enumerate(mapping_headers, 1):
    cell = mapping.cell(1, col, value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border
mapping_rows = [
    [1, "DoIP", "DoIP_ECU", "TCP:13400", "TCP", "192.168.0.100", 13400, "0x1008", "DUT DoIP Gateway", "TODO", "", "", "", "是", "修改为实际 ECU 地址"],
    [2, "Modbus", "Modbus_PLC", "TCP:502", "TCP", "192.168.0.101", 502, 1, "PLC/Device", "TODO", "", "", "", "是", "修改为实际从站和寄存器映射"],
    [3, "SCPI", "SCPI_DMM", "TCP:5025", "TCP", "192.168.0.102", 5025, "", "DMM", "TODO", "", "", "", "是", "确认仪器已开启远程控制"],
]
for row, values in enumerate(mapping_rows, 2):
    for col, value in enumerate(values, 1):
        cell = mapping.cell(row, col, value if value != "" else None)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
for col, width in enumerate([10, 12, 18, 14, 12, 18, 12, 12, 22, 16, 10, 10, 10, 10, 34], 1):
    mapping.column_dimensions[chr(64 + col)].width = width
mapping.freeze_panes = "A2"

trace = wb.create_sheet("产品追溯信息表")
trace.append(["字段", "值", "说明"])
for cell in trace[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = border
trace_rows = [
    ["产品型号", "TODO", "执行前填写 DUT 型号"],
    ["产品序列号", "TODO", "执行前填写 DUT SN"],
    ["测试用例版本", "1.0.0", "模板版本"],
    ["操作人员", "TODO", "执行前填写"],
    ["测试日期", "TODO", "执行前填写 YYYY-MM-DD"],
    ["设备配置文件", "Workflows/HardwareConfigs/real-device-acceptance.example.json", "对应 JSON 配置"],
    ["执行前确认", "已替换所有 TODO、IP、地址和上下限", "未确认不得连接真实设备"],
]
for row in trace_rows:
    trace.append(row)
for row in trace.iter_rows():
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
for col, width in enumerate([22, 70, 50], 1):
    trace.column_dimensions[chr(64 + col)].width = width
trace.freeze_panes = "A2"

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)
print(OUTPUT)
