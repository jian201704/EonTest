"""Generate a sample test case .xlsx file for EonTest CELL demo."""

import os

try:
    from openpyxl import Workbook
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl -q")
    from openpyxl import Workbook

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_test_xlsx(path):
    wb = Workbook()

    # ── Sheet 1: 测试工步表 ──
    ws1 = wb.active
    ws1.title = "测试工步表"

    headers = ["工步号", "设备分类", "设备名称", "通道/端口号", "测试项名称",
               "动作指令", "配置参数1", "配置参数2", "配置参数3",
               "下限值", "上限值", "单位", "超时(ms)", "失败处理", "备注"]

    # style
    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3050")
    thin_border = Border(
        left=Side(style='thin', color='2a3a5e'),
        right=Side(style='thin', color='2a3a5e'),
        top=Side(style='thin', color='2a3a5e'),
        bottom=Side(style='thin', color='2a3a5e')
    )

    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # test steps data
    tests = [
        # [stepNo, device, name, channel, testItem, action, param1, param2, param3, low, high, unit, timeout, failPolicy, remark]
        [1, "电源", "DMM Power Supply", "COM3", "Power ON",
         "power_on", "5.0V", "1.0A", "", "", "", "", 3000, "停机", "上电DUT"],
        [2, "万用表", "Digital Multimeter", "GPIB::1", "Measure 3.3V",
         "read_volt", "", "", "", "3.15", "3.45", "V", 5000, "重试1次", "3.3V电源轨"],
        [3, "万用表", "Digital Multimeter", "GPIB::1", "Measure 5.0V",
         "read_volt", "", "", "", "4.75", "5.25", "V", 5000, "停机", "5.0V电源轨"],
        [4, "万用表", "Digital Multimeter", "GPIB::1", "Measure 12V",
         "read_volt", "", "", "", "11.40", "12.60", "V", 5000, "停机", "12V输入"],
        [5, "万用表", "Digital Multimeter", "GPIB::1", "Measure Current",
         "read_current", "", "", "", "0.01", "0.05", "A", 8000, "停机", "待机电流"],
        [6, "万用表", "Digital Multimeter", "GPIB::1", "Measure Resistance",
         "read_resistance", "", "", "", "9500", "10500", "Ω", 5000, "重试1次", "输入阻抗"],
        [7, "Sample", "Analyzer", "", "Result Analysis",
         "analyze", "", "", "", "95", "100", "%", 5000, "停机", "综合判定"],
        [8, "Sample", "CSV Reporter", "", "Generate Report",
         "report", "", "", "", "", "", "", 5000, "继续", "生成测试报告"],
        [9, "电源", "DMM Power Supply", "COM3", "Power OFF",
         "power_off", "", "", "", "", "", "", 2000, "继续", "下电DUT"],
    ]

    for r, row_data in enumerate(tests, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val if val != "" else None)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # set column widths
    widths = [8, 14, 22, 14, 18, 14, 10, 10, 10, 12, 12, 8, 10, 12, 16]
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[chr(64 + i) if i < 27 else 'A'].width = w

    # ── Sheet 2: 设备通道映射表 ──
    ws2 = wb.create_sheet("设备通道映射表")

    dev_headers = ["设备编号", "设备分类", "设备名称", "通道号", "通道别名",
                   "连接方式", "主地址", "从地址", "波特率", "数据位",
                   "停止位", "校验位", "超时(ms)", "启用"]

    for c, h in enumerate(dev_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    devices = [
        [1, "电源", "DMM Power Supply", "CH1", "PSU-5V",
         "串口", "COM3", "", "9600", "8", "1", "None", 3000, "是"],
        [2, "万用表", "Digital Multimeter", "GPIB1", "DMM-34461A",
         "GPIB", "1", "", "", "", "", "", 5000, "是"],
    ]

    for r, row_data in enumerate(devices, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val if val != "" else None)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    wb.save(path)
    print(f"✅ Test case created: {path}")
    print(f"   File size: {os.path.getsize(path)} bytes")


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Workflows", "DMM_Test_Case.xlsx")
    create_test_xlsx(out)
